import pandas as pd
from openpyxl import load_workbook

# Load the existing Excel sheet
existing_workbook = load_workbook('existing_sheet.xlsx')

# Select the sheet where you want to add points
existing_sheet = existing_workbook['Sheet1']

# Load the new data from the input Excel sheet
new_data = pd.read_excel('input_sheet.xlsx')

# Loop through the rows in the new data
for index, row in new_data.iterrows():
    # Get the name and points from the row
    name = row['Name']
    points = row['Points']
    
    # Find the row with the same name in the existing sheet
    existing_row = existing_sheet.find(name)
    
    # If the name is not found, add a new row with the name and points
    if not existing_row:
        new_row = [name, points]
        existing_sheet.append(new_row)
    else:
        # If the name is found, add the points to the existing row
        row_num = existing_row.row
        existing_sheet.cell(row=row_num, column=2).value += points

# Save the changes to the existing Excel sheet
existing_workbook.save('existing_sheet.xlsx')
